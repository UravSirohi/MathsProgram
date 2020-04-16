import openpyxl as xl
from openpyxl.chart import LineChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_call = sheet.cell(row, 4)
        corrected_price_call.value = corrected_price

    values = Reference(wb['Sheet1'], min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

    chart = LineChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)
    print('Done')


file = input('Name of file:')
process_workbook(file)
